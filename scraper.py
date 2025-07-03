from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import time

options = Options()
options.add_argument("--start-maximized")
driver_path = "chromedriver.exe"
service = Service(driver_path)

driver = webdriver.Chrome(service=service, options=options)
driver.get("https://sourcing.alibaba.com/rfq/rfq_search_list.htm?spm=a2700.8073608.1998677541.82be65aaoUUltC&country=AE&recently=Y&tracelog=newest")
time.sleep(5)

wb = Workbook()
ws = wb.active
ws.title = "output"
columns = ["TITLE", "DESCRIPTION", "QUANTITY", "COUNTRY", "QUOTES LEFT", "DATE POSTED"]
ws.append(columns)

for col in range(1, len(columns) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)

for page in range(1, 101):
    print(f"Scraping page {page}...")
    time.sleep(3)
    items = driver.find_elements(By.CSS_SELECTOR, ".brh-rfq-item__main-info")
    for item in items:
        try:
            title = item.find_element(By.CLASS_NAME, "brh-rfq-item__subject").text.strip()
            desc = item.find_element(By.CLASS_NAME, "brh-rfq-item__detail").text.strip()
            text = item.text.strip()
            parts = text.split("\n")
            quantity = country = quotes_left = date_posted = ""
            for part in parts:
                if "Quantity Required" in part:
                    quantity = part.split("Quantity Required:")[-1].strip()
                elif "Posted in:" in part:
                    country = part.split("Posted in:")[-1].strip()
                elif "Quotes Left" in part:
                    quotes_left = part.split("Quotes Left")[-1].strip()
                elif "minutes before" in part or "hours before" in part:
                    date_posted = part.strip()
            ws.append([title, desc, quantity, country, quotes_left, date_posted])
        except:
            continue

    if page < 100:
        try:
            time.sleep(2)
            page_buttons = driver.find_elements(By.XPATH, "//a[contains(@href, 'page=')]")
            found = False
            for btn in page_buttons:
                if btn.text.strip() == str(page + 1):
                    driver.execute_script("arguments[0].scrollIntoView();", btn)
                    driver.execute_script("arguments[0].click();", btn)
                    print(f"Clicked page {page + 1}")
                    found = True
                    break
            if not found:
                print(f"Next page ({page + 1}) button not found. Stopping.")
                break
        except Exception as e:
            print(f"Error clicking page {page + 1}: {e}")
            break

for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

wb.save("output.xlsx")
driver.quit()
print("✅ Done! Data saved in 'output.xlsx'")
